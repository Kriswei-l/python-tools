from  openpyxl import  Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import datetime
import os


class ImageSystem:
    """ 图片列表生成excel文件系统"""
    def __init__(self, dirname):
        self.dirpath = dirname

    image_list = []

    def listfile(self, path, img_type='.jpg'):
        '获取指定路径下所有的图片文件, 返回列表'
        img_File_List = os.listdir(path)  # 图片列表
        for filename in img_File_List:
            filepath = os.path.join(path, filename)  # 图片的绝对路径
            if os.path.isdir(filepath):
                self.listfile(filepath, img_type)
                # print(filepath)
            else:
                if os.path.isfile(filepath) and filename.lower().endswith(img_type):
                    # print(os.path.join(path, filename))
                    self.image_list.append(os.path.join(path, filename))
        return self.image_list
    
    def wirte_execl(self, image_list, index, keySheet, path=''):
        imgsize = (720 / 10, 1280 / 10)  # 设置一个图像缩小的比例
        wb = load_workbook(path)
        ws = wb[keySheet]
        
        ws['A'+str(index)] = image_list.replace(self.dirpath, '')
        ws.column_dimensions['A'].width = 90  # 修改列A的宽

        ws.column_dimensions['B'].width = imgsize[0] * 1.1  # 修改列A的宽

        img = Image(image_list)  # 缩放图片
        s = 1
        if img.width > imgsize[0] :
            s = imgsize[0] / img.width
        elif img.height > imgsize[1] :
            s = imgsize[1] / img.height

        img.width = img.width*s
        img.height = img.height*s

        ws.add_image(img, 'B'+str(index))  # 图片 插入 A1 的位置上
        ws.row_dimensions[index].height = imgsize[1] * 0.78  # 修改列第1行的宽
        wb.save(path)


    def xlsx_create(self, image_list, filename='Workbook.xlsx'):
        '创建excel表格'

        imgsize = (720 / 4, 1280 / 4)  # 设置一个图像缩小的比例
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '文件名'
        ws['B1'] = '图片'

        # for s in image_list:
            # ws.append([s, self.get_FileSize(s)])
        ws.append([datetime.datetime.now()])

        output = os.path.join('./', filename)
        print('创建excel表格')
        wb.save(output)

        for key,value in image_list.items():
            # print(key+':'+value)
            wb = load_workbook(output)
            wb.create_sheet(key)
            ws = wb[key]
            ws['A1'] = '文件名'
            ws['B1'] = '图片'
            wb.save(output)

            i = 1
            for s in value:
                i+=1
                self.wirte_execl(s, i, key, output)

    def get_FileSize(self, filename):
        '获取文件的大小,结果保留两位小数，单位为kb'
        fsize = os.path.getsize(filename)
        fsize = fsize/float(1024)
        return f'{round(fsize, 2)} kb'